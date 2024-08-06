import time
import requests
import pandas as pd
from requests.auth import HTTPBasicAuth
from bs4 import BeautifulSoup
from requests.exceptions import HTTPError, ConnectionError, Timeout
from datetime import datetime
import os
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Function to fetch all work item IDs for a project
def fetch_work_item_ids(organization_name, ado_org_url, project_name, pat):
    wiql_url = f"{ado_org_url}/{organization_name}/{project_name}/_apis/wit/wiql?api-version=6.0"
    wiql_query = {"query": f"SELECT [System.Id] FROM workitems WHERE [System.TeamProject] = '{project_name}'"}
    response = requests.post(wiql_url, json=wiql_query, auth=HTTPBasicAuth('', pat))
    response.raise_for_status()
    work_item_ids = [item['id'] for item in response.json()['workItems']]
    return work_item_ids

# Function to fetch work item details in batches with retries and rate limiting
def fetch_work_item_details(organization_name, ado_org_url, pat, work_item_ids, project_name):
    work_items_url = f"{ado_org_url}/{organization_name}/{project_name}/_apis/wit/workitemsbatch?api-version=6.0"
    chunk_size = 200  # Azure DevOps allows fetching up to 200 work items per batch
    work_item_details = []

    for i in range(0, len(work_item_ids), chunk_size):
        ids_chunk = work_item_ids[i:i + chunk_size]
        success = False
        retries = 3
        while not success and retries > 0:
            try:
                response = requests.post(
                    work_items_url,
                    json={"ids": ids_chunk},
                    headers={"Content-Type": "application/json"},
                    auth=HTTPBasicAuth('', pat),
                    timeout=30  # Increase timeout as needed
                )
                response.raise_for_status()
                work_item_details.extend(response.json().get('value', []))
                success = True
            except (HTTPError, ConnectionError, Timeout) as e:
                print(f"Error occurred: {e}")
                retries -= 1
                time.sleep(5)  # Wait before retrying
                if retries == 0:
                    print("Max retries reached. Moving to the next chunk.")
        time.sleep(1)  # Add delay to respect rate limits

    return work_item_details

# Function to extract relevant details and clean HTML tags from descriptions
def extract_details(work_item_details):
    data = []
    for item in work_item_details:
        fields = item['fields']
        description_html = fields.get('System.Description', '')
        description_text = BeautifulSoup(description_html, 'html.parser').get_text() if description_html else ''
        data.append({
            'ID': item['id'],
            'Key': fields.get('System.WorkItemType', '') + '-' + str(item['id']),
            'Summary': fields.get('System.Title', ''),
            'Description': description_text,
            'Assignee': fields.get('System.AssignedTo', {}).get('displayName', ''),
            'Reporter': fields.get('System.CreatedBy', {}).get('displayName', ''),
            'Issue Type': fields.get('System.WorkItemType', ''),
            'Time Estimate': fields.get('Microsoft.VSTS.Scheduling.OriginalEstimate', 0),
            'Time Spent': fields.get('Microsoft.VSTS.Scheduling.CompletedWork', 0),
            'Due Date': fields.get('Microsoft.VSTS.Scheduling.DueDate', ''),
            'Created Date': fields.get('System.CreatedDate', '')
        })
    return data

# Function to calculate statistics
def calculate_statistics(work_item_details):
    work_item_counts = {}
    for item in work_item_details:
        work_item_type = item['fields'].get('System.WorkItemType', '')
        work_item_counts[work_item_type] = work_item_counts.get(work_item_type, 0) + 1
    return work_item_counts

# Function to auto-fit columns
def auto_fit_columns(ws):
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

# Function to add a report details sheet
def add_report_details_sheet(excel_writer, start_time, input_parameters, total_issues):
    report_title = "ADO Discovery Report"
    purpose_of_report = "This report gives all the work item details and statistics"
    run_date = start_time.strftime("%d-%b-%Y %I:%M %p")
    run_duration = str(datetime.now() - start_time)
    run_by = os.getlogin()
    discovery_status = "Completed"

    report_details = {
        "Report Title": report_title,
        "Purpose of Report": purpose_of_report,
        "Run Date": run_date,
        "Run Duration": run_duration,
        "Run By": run_by,
        "Total Issue Count": total_issues,
        "Discovery Status": discovery_status,
        "Project Name": input_parameters.get('project_name', ''),
        "Base URL": input_parameters.get('ado_org_url', ''),
        "Username": input_parameters.get('username', ''),
        "Project Key": input_parameters.get('project_key', '')
    }

    report_df = pd.DataFrame(list(report_details.items()), columns=['Parameter', 'Value'])
    report_df.to_excel(excel_writer, index=False, sheet_name='Report Details')

    wb = excel_writer.book
    ws = wb['Report Details']
    for cell in ws["A"]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(fill_type='solid', start_color='252D76', end_color='252D76')
        cell.alignment = Alignment(horizontal='left')
    
    for cell in ws["B"]:
        cell.alignment = Alignment(horizontal='left')

    auto_fit_columns(ws)

# Function to format the Excel sheets
def format_excel_sheets(excel_writer):
    wb = excel_writer.book
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Set font and alignment for headers
        for cell in ws[1]:
            cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
            cell.fill = PatternFill(fill_type='solid', start_color='252D76', end_color='252D76')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Set font and alignment for data cells
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = Font(name='Calibri', size=11)
                if isinstance(cell.value, (int, float, datetime)):
                    cell.alignment = Alignment(horizontal='right')
                else:
                    cell.alignment = Alignment(horizontal='left')

        # Freeze panes
        ws.freeze_panes = ws['B2']

        # Auto-fit columns
        auto_fit_columns(ws)

# Function to process a single project
def process_project(project_row):
    # Strip any leading/trailing whitespace from column headers
    project_row = project_row.rename(index=lambda x: x.strip())

    organization_name = project_row['Organization Name']
    ado_org_url = project_row['Organization URL'].rstrip('/')
    project_name = project_row['Project Name']
    pat = project_row['PAT']
    username = project_row['Username']

    start_time = datetime.now()

    work_item_ids = fetch_work_item_ids(organization_name, ado_org_url, project_name, pat)
    work_item_details = fetch_work_item_details(organization_name, ado_org_url, pat, work_item_ids, project_name)
    work_item_data = extract_details(work_item_details)
    statistics = calculate_statistics(work_item_details)

    df = pd.DataFrame(work_item_data)
    excel_writer = pd.ExcelWriter(f'ado_test_files/{project_name}_details.xlsx', engine='openpyxl')

    # Add the report details sheet first
    add_report_details_sheet(excel_writer, start_time, {
        'ado_org_url': ado_org_url,
        'project_name': project_name,
        'username': username,
        'project_key': organization_name  # Assuming 'Organization Name' is used as 'Project Key'
    }, len(work_item_ids))

    df.to_excel(excel_writer, index=False, sheet_name='Work Item Details')
    statistics_df = pd.DataFrame.from_dict(statistics, orient='index', columns=['Count'])
    statistics_df.index.name = 'Work Item Type'
    statistics_df.to_excel(excel_writer, sheet_name='Statistics')

    # Apply formatting to all sheets
    format_excel_sheets(excel_writer)

    excel_writer._save()

projects_df = pd.read_excel('ado_project_details.xlsx')

for index, project_row in projects_df.iterrows():
    process_project(project_row)

print("Work item details and statistics saved to Excel files.")
