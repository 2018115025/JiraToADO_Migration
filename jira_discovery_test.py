import requests
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from collections import Counter
from datetime import datetime
import os

# Function to fetch JIRA issues from a project with pagination
def fetch_jira_issues(base_url, username, api_token, project_key):
    session = requests.Session()
    session.auth = (username, api_token)

    url = f"{base_url}/rest/api/3/search"
    jql_query = f"project = {project_key}"
    max_results = 100
    start_at = 0
    all_issues = []

    while True:
        params = {
            'jql': jql_query,
            'fields': 'key,summary,description,assignee,reporter,issuetype,timeZone,timeestimate,timespent,duedate,created',
            'startAt': start_at,
            'maxResults': max_results
        }

        response = session.get(url, params=params)
        response.raise_for_status()
        

        data = response.json()
        issues = data.get('issues', [])
        all_issues.extend(issues)

        if len(issues) < max_results:
            break

        start_at += max_results

    return all_issues

# Function to extract text content from description field
def extract_description(description_data):
    if not description_data:
        return 'no description'

    description_text = ''
    for content in description_data['content']:
        if content['type'] == 'paragraph':
            for text in content['content']:
                if text['type'] == 'text':
                    description_text += text['text']
        elif content['type'] == 'table':
            for row in content['content']:
                for cell in row['content']:
                    for text in cell['content']:
                        if text['type'] == 'text':
                            description_text += text['text']
    return description_text

# Function to add a report details sheet
def add_report_details_sheet(workbook, start_time, input_parameters, total_issues):
    report_title = "Jira Discovery Report"
    purpose_of_report = "This report gives all the work item details and statistics"
    run_date = start_time.strftime("%d-%b-%Y %I:%M %p")
    run_duration = str(datetime.now() - start_time)
    run_by = os.getlogin()
    discovery_status = "Completed"

    report_sheet = workbook.create_sheet(title="Report Details", index=0)
    headers = ["Report Title", "Purpose of Report", "Run Date", "Run Duration", "Run By", "Total Issue Count", "Discovery Status"]
    for row_num, header in enumerate(headers, 1):
        cell = report_sheet.cell(row=row_num, column=1, value=header)
        cell.font = Font(bold=True, color='FFFFFF', name='Calibri', size=11)
        cell.fill = PatternFill(fill_type='solid', start_color='252D76', end_color='252D76')
        cell.alignment = Alignment(horizontal='left')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    report_details = [report_title, purpose_of_report, run_date, run_duration, run_by, total_issues, discovery_status]

    for row_num, row in enumerate(report_details, 1):
        cell = report_sheet.cell(row=row_num, column=2, value=row)
        cell.alignment = Alignment(horizontal='left')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    row_num = len(headers) + 1
    param_row = row_num
    value_row = row_num
    for param, value in input_parameters.items():
        cell = report_sheet.cell(row=param_row, column=1, value=param[0].upper() + param[1:])
        cell.font = Font(bold=True, color='FFFFFF', name='Calibri', size=11)
        cell.fill = PatternFill(fill_type='solid', start_color='252D76', end_color='252D76')
        cell.alignment = Alignment(horizontal='left')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        param_row += 1

        cell = report_sheet.cell(row=value_row, column=2, value=value)
        cell.alignment = Alignment(horizontal='left')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        value_row += 1

    # Auto-fit columns for better readability
    for col in range(1, 3):
        column_width = max(len(str(report_sheet.cell(row=row, column=col).value)) for row in range(1, row_num + 1))
        report_sheet.column_dimensions[get_column_letter(col)].width = column_width + 2

# Function to format the Excel sheet
def format_excel_sheet(sheet):
    # Set header style
    for col_num, cell in enumerate(sheet[1], 1):
        cell.font = Font(bold=True, color='FFFFFF', name='Calibri', size=11)
        cell.fill = PatternFill(fill_type='solid', start_color='252D76', end_color='252D76')
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Freeze the top row
    sheet.freeze_panes = sheet['A2']

    # Set column widths and cell alignment
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width

        for cell in col:
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal='right')
            elif isinstance(cell.value, datetime):
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.alignment = Alignment(horizontal='left')

# Function to generate a report for a project
def generate_project_report(base_url, username, api_token, project_key, project_name):
    start_time = datetime.now()

    # Fetch JIRA issues for the project
    jira_data = fetch_jira_issues(base_url, username, api_token, project_key)
    total_issues = len(jira_data)

    # Create a new Excel workbook
    wb = openpyxl.Workbook()
    
    # Sheet 1: Detailed JIRA Issues
    ws1 = wb.active
    ws1.title = f"JIRA Report - {project_name}"

    # Define headers for the Excel sheet
    headers = ["S.No", "Project Name", "Issue Key", "Summary", "Description", "Assignee", "Reporter", "Issue Type", "Time Zone",
               "Estimated Hours", "Hours Worked", "Due Date", "Date Assigned"]
    for col_num, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col_num, value=header)

    issue_types = Counter()
    row_number = 2
    for idx, issue in enumerate(jira_data, start=1):
        issue_key = issue['key']
        summary = issue['fields']['summary']
        description = extract_description(issue['fields'].get('description'))
        assignee = issue['fields']['assignee']['displayName'] if issue['fields']['assignee'] else 'Unassigned'
        reporter = issue['fields']['reporter']['displayName'] if issue['fields']['reporter'] else 'Unknown'
        issue_type = issue['fields']['issuetype']['name']
        timeZone = issue['fields']['reporter']['timeZone'] if issue['fields'].get('reporter') and issue['fields']['reporter'].get('timeZone') else 'Unknown'
        estimated_hours = issue['fields'].get('timeestimate', 0) / 3600 if issue['fields'].get('timeestimate') else 0
        hours_worked = issue['fields'].get('timespent', 0) / 3600 if issue['fields'].get('timespent') else 0
        due_date = issue['fields']['duedate'] if issue['fields'].get('duedate') else 'No due date'
        date_assigned = datetime.strptime(issue['fields']['created'], "%Y-%m-%dT%H:%M:%S.%f%z").strftime("%Y-%m-%d %H:%M:%S")

        # Count issue types
        issue_types[issue_type] += 1

        # Add data to the Excel row
        ws1.cell(row=row_number, column=1, value=idx)
        ws1.cell(row=row_number, column=2, value=project_name)
        ws1.cell(row=row_number, column=3, value=issue_key)
        ws1.cell(row=row_number, column=4, value=summary)
        ws1.cell(row=row_number, column=5, value=description)
        ws1.cell(row=row_number, column=6, value=assignee)
        ws1.cell(row=row_number, column=7, value=reporter)
        ws1.cell(row=row_number, column=8, value=issue_type)
        ws1.cell(row=row_number, column=9, value=timeZone)
        ws1.cell(row=row_number, column=10, value=estimated_hours)
        ws1.cell(row=row_number, column=11, value=hours_worked)
        ws1.cell(row=row_number, column=12, value=due_date)
        ws1.cell(row=row_number, column=13, value=date_assigned)

        row_number += 1

    # Format the Detailed JIRA Issues sheet
    format_excel_sheet(ws1)

    # Sheet 2: Statistics
    ws2 = wb.create_sheet(title="Statistics")
    ws2.append(["Work Item Type", "Count"])
    for issue_type, count in issue_types.items():
        ws2.append([issue_type, count])

    # Format the Statistics sheet
    format_excel_sheet(ws2)

    # Add report details sheet
    input_parameters = {
        'base_url': base_url,
        'username': username,
        'project_key': project_key,
        'project_name': project_name
    }
    add_report_details_sheet(wb, start_time, input_parameters, total_issues)

    # Format the Report Details sheet
    format_excel_sheet(wb['Report Details'])

    # Save the Excel workbook
    wb.save(f"jira_test_files/jira_report_{project_name}.xlsx")

    print(f"Excel report generated: jira_report_{project_name}.xlsx")

# Load project details from an Excel sheet
def load_project_details(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    projects = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        projects.append({
            'name': row[0],
            'baseurl': row[1],
            'key': row[2],
            'username': row[3],
            'api_token': row[4]
        })

    return projects

# Main function
def main():
    # Load project details from Excel
    project_details = load_project_details('jira_test_details.xlsx')

    # Generate report for each project
    for project in project_details:
        generate_project_report(project['baseurl'], project['username'], project['api_token'], project['key'], project['name'])

if __name__ == "__main__":
    main()
