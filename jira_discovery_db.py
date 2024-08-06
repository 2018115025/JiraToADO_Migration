import requests
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from collections import Counter
from datetime import datetime
import psycopg2

# Function to fetch JIRA issues from a project with pagination
def fetch_jira_issues(base_url, username, api_token, project_key):
    session = requests.Session()
    session.auth = (username, api_token)

    # API endpoint to fetch issues for the project
    url = f"{base_url}/rest/api/3/search"
    jql_query = f"project = {project_key}"
    max_results = 100
    start_at = 0
    all_issues = []

    while True:
        params = {
            'jql': jql_query,
            'fields': 'key,summary,description,assignee,reporter,issuetype,timeZone,timeestimate,timespent,duedate,created',
            'startAt': start_at,
            'maxResults': max_results
        }

        response = session.get(url, params=params)
        response.raise_for_status()

        data = response.json()
        issues = data.get('issues', [])
        all_issues.extend(issues)

        if len(issues) < max_results:
            break

        start_at += max_results

    return all_issues

# Function to extract text content from description field
def extract_description(description_data):
    if not description_data:
        return 'no description'

    description_text = ''
    for content in description_data['content']:
        if content['type'] == 'paragraph':
            for text in content['content']:
                if text['type'] == 'text':
                    description_text += text['text']
        elif content['type'] == 'table':
            for row in content['content']:
                for cell in row['content']:
                    for text in cell['content']:
                        if text['type'] == 'text':
                            description_text += text['text']
    return description_text

# Function to connect to PostgreSQL
def connect_to_db():
    conn = psycopg2.connect(
        dbname="jiratoadodb",
        user="postgres",
        password="0000",
        host="40.117.145.14",
        port="5432"
    )
    print("Connected to the database")
    return conn

# Function to store JIRA issues in PostgreSQL
def store_jira_issues(conn, project_name, jira_data):
    cursor = conn.cursor()

    insert_query = """
    INSERT INTO jira_issue_details (
        project_name, issue_key, summary, description, assignee, reporter, issue_type, timezone, estimated_hours, hours_worked, due_date, date_assigned
    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    """

    for issue in jira_data:
        issue_key = issue['key']
        summary = issue['fields']['summary']
        description = extract_description(issue['fields'].get('description'))
        assignee = issue['fields']['assignee']['displayName'] if issue['fields']['assignee'] else 'Unassigned'
        reporter = issue['fields']['reporter']['displayName'] if issue['fields']['reporter'] else 'Unknown'
        issue_type = issue['fields']['issuetype']['name']
        timezone = issue['fields']['reporter']['timeZone'] if issue['fields'].get('reporter') and issue['fields']['reporter'].get('timeZone') else 'Unknown'
        estimated_hours = issue['fields'].get('timeestimate', 0) / 3600 if issue['fields'].get('timeestimate') else 0
        hours_worked = issue['fields'].get('timespent', 0) / 3600 if issue['fields'].get('timespent') else 0
        due_date = issue['fields']['duedate'] if issue['fields'].get('duedate') else None
        date_assigned = datetime.strptime(issue['fields']['created'], "%Y-%m-%dT%H:%M:%S.%f%z").strftime("%Y-%m-%d %H:%M:%S")

        cursor.execute(insert_query, (
            project_name, issue_key, summary, description, assignee, reporter,
            issue_type, timezone, estimated_hours, hours_worked, due_date, date_assigned
        ))

        print(f"Task {issue_key} was added")

    conn.commit()
    print("All JIRA details added")
    cursor.close()

# Function to store issue statistics in PostgreSQL
def store_issue_statistics(conn, project_name, issue_types):
    cursor = conn.cursor()

    insert_query = """
    INSERT INTO jira_issue_stats (project_name, issue_type, count) VALUES (%s, %s, %s)
    """

    for issue_type, count in issue_types.items():
        cursor.execute(insert_query, (project_name, issue_type, count))
        print(f"Statistics for project {project_name}, issue type {issue_type} with count {count} was added")

    conn.commit()
    print("All JIRA stats added")
    cursor.close()

# Function to generate a report for a project and store it in PostgreSQL
def generate_project_report(base_url, username, api_token, project_key, project_name, conn):
    # Fetch JIRA issues for the project
    jira_data = fetch_jira_issues(base_url, username, api_token, project_key)

    # Store the fetched JIRA issues in PostgreSQL
    store_jira_issues(conn, project_name, jira_data)

    # Count issue types for statistics
    issue_types = Counter(issue['fields']['issuetype']['name'] for issue in jira_data)

    # Store issue statistics in PostgreSQL
    store_issue_statistics(conn, project_name, issue_types)

    print(f"JIRA data for project {project_name} has been stored in the database.")

# Load project details from an Excel sheet
def load_project_details(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    projects = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        projects.append({
            'name': row[0],
            'baseurl': row[1],
            'key': row[2],
            'username': row[3],
            'api_token': row[4]
        })

    return projects

# Main function
def main():
    # Load project details from Excel
    project_details = load_project_details('jira_project_details_db.xlsx')

    # Connect to the PostgreSQL database
    conn = connect_to_db()

    # Generate report and store data for each project
    for project in project_details:
        generate_project_report(project['baseurl'], project['username'], project['api_token'], project['key'], project['name'], conn)

    # Close the database connection
    conn.close()

if __name__ == "__main__":
    main()
