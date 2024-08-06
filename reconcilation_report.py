import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from datetime import datetime
import os

def load_excel_data(file_path):
    # Load the data from the Excel file
    xls = pd.ExcelFile(file_path)
    sheets = xls.sheet_names
    data = {sheet: xls.parse(sheet) for sheet in sheets}
    return data

def compare_statistics(jira_stats, ado_stats):
    # Map JIRA Story to ADO User Story
    jira_stats['Work Item Type'] = jira_stats['Work Item Type'].replace('Story', 'User Story')
    
    # Sum Task and Sub-task counts in JIRA and combine them under 'Task'
    jira_stats['Work Item Type'] = jira_stats['Work Item Type'].replace({'Sub-task': 'Task'})
    jira_task_counts = jira_stats.groupby('Work Item Type')['Count'].sum().to_frame()
    
    # Map ADO statistics similarly
    ado_task_counts = ado_stats.groupby('Work Item Type')['Count'].sum().to_frame()
    
    # Combine the counts for comparison
    comparison = pd.merge(jira_task_counts, ado_task_counts, left_index=True, right_index=True, how='outer', suffixes=('_JIRA', '_ADO'))
    comparison = comparison.fillna(0).astype(int)
    
    # Determine if counts match
    comparison['Status'] = comparison['Count_JIRA'] == comparison['Count_ADO']
    comparison['Status'] = comparison['Status'].map({True: 'Matched', False: 'Unmatched'})
    
    # Include both JIRA and ADO work item types
    comparison = comparison.reset_index()
    comparison['JIRA Work Item Type'] = comparison['Work Item Type']
    comparison['ADO Work Item Type'] = comparison['Work Item Type']
    
    return comparison

def create_report_details_sheet(wb, start_time, total_issues, matched_issues, unmatched_issues, Jira_Project, ADO_project):
    ws = wb.create_sheet(title="Report Details", index=0)
    report_title = "Reconciliation Report"
    purpose_of_report = "This report reconciles work item counts between JIRA and ADO."
    run_date = start_time.strftime("%d-%b-%Y %I:%M %p")
    run_duration = str(datetime.now() - start_time)
    run_by = os.getlogin()

    report_details = [
        ("Report Title", report_title),
        ("Purpose of Report", purpose_of_report),
        ("Run Date", run_date),
        ("Run Duration", run_duration),
        ("Run By", run_by),
        ("Jira Project Name", Jira_Project),
        ("ADO Project Name", ADO_project),
        ("Total Issues", total_issues),
        ("Matched Issues", matched_issues),
        ("Unmatched Issues", unmatched_issues)
    ]

    for row_num, (param, value) in enumerate(report_details, 1):
        ws.cell(row=row_num, column=1, value=param)
        ws.cell(row=row_num, column=2, value=value)

        for cell in ws[row_num]:
            cell.font = Font(bold=True, color='FFFFFF') if cell.column == 1 else Font(bold=False)
            cell.fill = PatternFill(fill_type='solid', start_color='252D76', end_color='252D76') if cell.column == 1 else PatternFill(fill_type='solid', start_color='FFFFFF', end_color='FFFFFF')
            cell.alignment = Alignment(horizontal='left')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

def auto_fit_columns(ws):
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

def format_excel_sheet(sheet):
    # Set header style
    for col_num, cell in enumerate(sheet[1], 1):
        cell.font = Font(bold=True, color='FFFFFF', name='Calibri', size=11)
        cell.fill = PatternFill(fill_type='solid', start_color='252D76', end_color='252D76')
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Freeze the top row
    sheet.freeze_panes = sheet['A2']

    # Set column widths and cell alignment
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width

        for cell in col:
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal='right')
            elif isinstance(cell.value, datetime):
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.alignment = Alignment(horizontal='left')

def create_reconciliation_report(jira_file_path, ado_file_path, project_name, Jira_Project, ADO_project):
    start_time = datetime.now()

    # Load data
    jira_data = load_excel_data(jira_file_path)
    ado_data = load_excel_data(ado_file_path)

    # Extract statistics
    jira_stats = jira_data['Statistics']
    ado_stats = ado_data['Statistics']
    
    # Compare statistics
    comparison = compare_statistics(jira_stats, ado_stats)
    matched_issues = comparison['Status'].value_counts().get('Matched', 0)
    unmatched_issues = comparison['Status'].value_counts().get('Unmatched', 0)
    total_issues = len(comparison)

    # Create reconciliation report
    wb = Workbook()
    ws = wb.active
    ws.title = "Reconciliation"
    
    # Add headers
    headers = ['JIRA Work Item Type', 'ADO Work Item Type', 'JIRA Count', 'ADO Count', 'Status']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, color='FFFFFF', name='Calibri', size=11)
        cell.fill = PatternFill(fill_type='solid', start_color='252D76', end_color='252D76')
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Add data rows
    for r_idx, row in enumerate(comparison.itertuples(index=False), 2):
        ws.cell(row=r_idx, column=1, value=row[4])
        ws.cell(row=r_idx, column=2, value=row[5])
        ws.cell(row=r_idx, column=3, value=row[1])
        ws.cell(row=r_idx, column=4, value=row[2])
        status_cell = ws.cell(row=r_idx, column=5, value=row[3])
        if row[3] == "Matched":
            status_cell.fill = PatternFill(start_color='009900', end_color='009900', fill_type='solid')
        else:
            status_cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    
    # Add total row
    total_jira = comparison['Count_JIRA'].sum()
    total_ado = comparison['Count_ADO'].sum()
    total_status = "Matched" if total_jira == total_ado else "Unmatched"

    total_row = len(comparison) + 2
    ws.cell(row=total_row, column=1, value="Total")
    ws.cell(row=total_row, column=2, value="")
    ws.cell(row=total_row, column=3, value=total_jira)
    ws.cell(row=total_row, column=4, value=total_ado)
    status_cell = ws.cell(row=total_row, column=5, value=total_status)

    # Apply conditional formatting for total status
    if total_status == "Matched":
        status_cell.fill = PatternFill(start_color='009900', end_color='009900', fill_type='solid')
    else:
        status_cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

    # Auto-fit columns
    auto_fit_columns(ws)

    # Format the Reconciliation sheet
    format_excel_sheet(ws)

    # Add report details sheet
    create_report_details_sheet(wb, start_time, total_issues, matched_issues, unmatched_issues, Jira_Project, ADO_project)

    # Auto-fit columns in all sheets
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        auto_fit_columns(ws)

    # Save the reconciliation report
    output_file_path = f'reconciliation/{project_name}_reconciliation_report.xlsx'
    wb.save(output_file_path)
    print(f"Reconciliation report saved to {output_file_path}")

# Define the paths to your Excel files
jira_file_path = 'jira_test_files/jira_report_Newt CAMPP.xlsx'
ado_file_path = 'ado_test_files/Newt CAMPP1_details.xlsx'
project_name = 'Newt CAMPP'
Jira_Project = 'Newt CAMPP'
ADO_project = 'Newt CAMPP1'

# Create the reconciliation report
create_reconciliation_report(jira_file_path, ado_file_path, project_name, Jira_Project, ADO_project)
